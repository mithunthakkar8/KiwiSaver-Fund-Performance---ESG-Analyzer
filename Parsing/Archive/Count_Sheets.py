from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook(r"C:\Users\mithu\Documents\MEGA\Projects\KiwiSaver Fund Performance & ESG Analyzer\Downloads\Milford_Asset\Investment_Funds_Fact_Sheets\Equity_Holdings_Data.xlsx", read_only=True)

# Count the number of sheets
sheet_count = len(wb.sheetnames)

print(f"Number of sheets: {sheet_count}")

# print(wb.sheetnames)
